from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.edge.options import Options
from selenium.webdriver.support.ui import Select, WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import pandas as pd
import time
import datetime
from openpyxl import load_workbook

class auto_run:
    def __init__(self, download_dir, input_excel_path=None, output_excel_path=None):
        self.download_dir = download_dir
        self.input_excel_path = input_excel_path
        self.output_excel_path = output_excel_path
        chrome_options = webdriver.ChromeOptions()
        chrome_options.add_experimental_option("prefs", {"download.default_directory": download_dir})
        #chrome_options.add_experimental_option('detach', True)
        self.driver = webdriver.Chrome(chrome_options)
        self.driver.get("https://apps.cra-arc.gc.ca/ebci/rhpd/beta/entry")
        self.__find_element('/html/body/app-root/rccr-wet-template/div/div/main/app-entry/form/rccr-button-group/div/fieldset/div[2]/rccr-button-toggle[1]/label/span').click() #click salary
        self.__find_element('/html/body/app-root/rccr-wet-template/div/div/main/app-entry/form/div[2]/div/button').click() #click next
        self.__run_from_excel()
        self.driver.quit()

    def __find_element(self,xpath):
        for i in range(0,10):
            try:
                element = WebDriverWait(self.driver,3).until(EC.presence_of_element_located((By.XPATH, xpath)))
            except:
                print(i)
                pass
            else:
                return element
        raise TimeoutError('reach max try, check the Internet connection')
            
    def __calculate(self, employee_name, employer_name, employment_province, pay_frequency, year, month, day, income_per_pay):
        self.__find_element('/html/body/app-root/rccr-wet-template/div/div/main/app-step1/form/rccr-text-input[1]/div/div/input').send_keys(employee_name) #input employee name
        self.__find_element('/html/body/app-root/rccr-wet-template/div/div/main/app-step1/form/rccr-text-input[2]/div/div/input').send_keys(employer_name) #input employer name
        Select(self.__find_element('/html/body/app-root/rccr-wet-template/div/div/main/app-step1/form/rccr-select[1]/div/select')).select_by_visible_text(employment_province) #select employment province
        Select(self.__find_element('/html/body/app-root/rccr-wet-template/div/div/main/app-step1/form/rccr-select[2]/div/select')).select_by_value(pay_frequency) #select pay frenquency
        Select(self.__find_element('/html/body/app-root/rccr-wet-template/div/div/main/app-step1/form/fieldset/div[2]/div/div/rccr-select[1]/div/select')).select_by_visible_text(year) #select date the employee is paid: year
        Select(self.__find_element('/html/body/app-root/rccr-wet-template/div/div/main/app-step1/form/fieldset/div[2]/div/div/rccr-select[2]/div/select')).select_by_visible_text(month) #select date the employee is paid: month
        Select(self.__find_element('/html/body/app-root/rccr-wet-template/div/div/main/app-step1/form/fieldset/div[2]/div/div/rccr-select[3]/div/select')).select_by_visible_text(day) #select date the employee is paid: day
        self.__find_element('/html/body/app-root/rccr-wet-template/div/div/main/app-step1/form/div/div/div/div/div/button').click() #click next
        self.__find_element('/html/body/app-root/rccr-wet-template/div/div/main/app-step2/form/fieldset[1]/rccr-currency-input[1]/div/div/input').send_keys(income_per_pay) #input income per pay period
        self.__find_element('/html/body/app-root/rccr-wet-template/div/div/main/app-step2/form/div[2]/div/div/div/div/button').click() #click next
        self.__find_element('/html/body/app-root/rccr-wet-template/div/div/main/app-step3/form/div/div/div/div/div/button').click() #click calulate
        self.__find_element('/html/body/app-root/rccr-wet-template/div/div/main/app-results/div/div[2]/button[1]').click() #click view in pdf
        #print('download at:', self.download_dir)
        cash_income = self.__find_element('/html/body/app-root/rccr-wet-template/div/div/main/app-results/app-results-salary/table/tbody/tr[2]/td[4]/strong').text #get total cash income
        federal_deduction = self.__find_element('/html/body/app-root/rccr-wet-template/div/div/main/app-results/app-results-salary/table/tbody/tr[3]/td[2]').text #get federal tax deduction
        provincial_deduction = self.__find_element('/html/body/app-root/rccr-wet-template/div/div/main/app-results/app-results-salary/table/tbody/tr[4]/td[2]').text #get provincial tax income
        CPP = self.__find_element('/html/body/app-root/rccr-wet-template/div/div/main/app-results/app-results-salary/table/tbody/tr[6]/td[3]').text #get CPP deduction
        if (int(year)>=2024): #2024 new tax law: CPP2
            EI = self.__find_element('/html/body/app-root/rccr-wet-template/div/div/main/app-results/app-results-salary/table/tbody/tr[8]/td[3]').text #get EI deduction
            total_deduction = self.__find_element('/html/body/app-root/rccr-wet-template/div/div/main/app-results/app-results-salary/table/tbody/tr[9]/td[4]').text #get total deduction
            net_amount = self.__find_element('/html/body/app-root/rccr-wet-template/div/div/main/app-results/app-results-salary/table/tbody/tr[10]/td[4]/strong').text # get net amount
        else:
            EI = self.__find_element('/html/body/app-root/rccr-wet-template/div/div/main/app-results/app-results-salary/table/tbody/tr[7]/td[3]').text #get EI deduction
            total_deduction = self.__find_element('/html/body/app-root/rccr-wet-template/div/div/main/app-results/app-results-salary/table/tbody/tr[8]/td[4]').text #get total deduction
            net_amount = self.__find_element('/html/body/app-root/rccr-wet-template/div/div/main/app-results/app-results-salary/table/tbody/tr[9]/td[4]/strong').text # get net amount
        self.__find_element('/html/body/app-root/rccr-wet-template/div/div/main/app-results/div/div[3]/button[2]').click() #click next calculation
        return {
                'cash_income':float(cash_income), 
                'provincial_deduction': float(provincial_deduction),
                'federal_deduction': float(federal_deduction), 
                'CPP': float(CPP), 
                'EI': float(EI), 
                'total_deduction': float(total_deduction), 
                'net_amount': float(net_amount)
            }


    def __run_from_excel(self):
        wb = load_workbook(self.input_excel_path)
        sheet = wb['员工工资信息表格']
        i=4
        while(1):
            if (sheet.cell(i,1).value==None):
                break
            employer_name = 'IRC'
            employee_name = sheet.cell(i,2).value
            employment_province = sheet.cell(i,6).value
            pay_frequency = 'MONTHLY_12PP'
            year = str(pd.to_datetime(sheet.cell(i,5).value).date().year)
            all_month = ['January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September', 'October', 'November', 'December']
            month = all_month[int(pd.to_datetime(sheet.cell(i,5).value).date().month)-1]
            day = int(pd.to_datetime(sheet.cell(i,5).value).date().day)
            if(day<10):
                day = '0'+str(day)
            else:
                day = str(day)
            income_per_pay = str(sheet.cell(i,7).value)
            result = self.__calculate(employee_name, employer_name, employment_province, pay_frequency, year, month, day, income_per_pay)
            sheet.cell(i,13).value = result['cash_income']
            sheet.cell(i,14).value = result['provincial_deduction']
            sheet.cell(i,15).value = result['federal_deduction']
            sheet.cell(i,16).value = result['CPP']
            sheet.cell(i,17).value = result['EI']
            sheet.cell(i,18).value = result['total_deduction']
            sheet.cell(i,19).value = result['net_amount']
            sheet['U1'] = str(datetime.datetime.now())
            i=i+1
        wb.save(self.output_excel_path) 